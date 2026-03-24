from openpyxl import load_workbook
import re

wb = load_workbook("file.xlsx")

from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

def col_letter(n):
    """1-indexed column number to letter"""
    result = ""
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result

def find_pct_columns(ws, header_row=1):
    """Return list of (col_idx, col_letter) for columns whose header contains % """
    pct_cols = []
    for cell in ws[header_row]:
        if cell.value and isinstance(cell.value, str) and '%' in cell.value.upper():
            pct_cols.append(cell.column)
    return pct_cols

def get_data_range(ws):
    """Return (min_row, max_row) for data (excluding header)"""
    rows = list(ws.iter_rows(values_only=True))
    return 2, len(rows)  # header is row 1

def apply_highlight(ws, header_row=1):
    pct_cols = find_pct_columns(ws, header_row)
    if not pct_cols:
        return 0
    
    min_row, max_row = get_data_range(ws)
    
    # First convert string % values like "35.22 %" to float in cells
    for row in ws.iter_rows(min_row=min_row, max_row=max_row):
        for cell in row:
            if cell.column in pct_cols and isinstance(cell.value, str):
                m = re.search(r'[\d.]+', cell.value)
                if m:
                    cell.value = float(m.group()) / 100  # store as 0-1 float
                    cell.number_format = '0.00%'
    
    for col_idx in pct_cols:
        cl = col_letter(col_idx)
        range_str = f"{cl}{min_row}:{cl}{max_row}"
        ws.conditional_formatting.add(range_str,
            CellIsRule(operator="greaterThan", formula=["0.8"], fill=red_fill))
    
    return len(pct_cols)

# --- DB sheet ---
ws_db = wb['DB']
apply_highlight(ws_db, header_row=1)

# --- APP sheet ---
ws_app = wb['APP']
apply_highlight(ws_app, header_row=1)

# --- Network sheet: CPU and RAM are in row 2 ---
ws_net = wb['Network']
# CPU col=7, RAM col=8 (1-indexed) based on header inspection
# Values are like "35%" strings - let's scan and convert
for row in ws_net.iter_rows(min_row=3):
    for cell in row:
        if cell.column in [7, 8] and isinstance(cell.value, str):
            m = re.search(r'[\d.]+', cell.value)
            if m:
                cell.value = float(m.group()) / 100
                cell.number_format = '0.00%'

max_row_net = ws_net.max_row
for col_idx in [7, 8]:
    cl = col_letter(col_idx)
    range_str = f"{cl}3:{cl}{max_row_net}"
    ws_net.conditional_formatting.add(range_str,
        CellIsRule(operator="greaterThan", formula=["0.8"], fill=red_fill))

wb.save("file_highlight.xlsx")
print("Done")
